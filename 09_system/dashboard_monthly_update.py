"""
糸貫グループ 月次ダッシュボード更新スクリプト
================================================
Excelファイルを読み込んで HTML を更新し、GitHub へ Push します。
Push 後に上司へ共有するURLを表示します。

使い方:
    pip install openpyxl
    python 09_system/dashboard_monthly_update.py

必要なもの:
    - 09_system/月次データ入力テンプレート.xlsx（記入済み）
    - git の認証設定（初回のみ）
"""

import re
import sys
import subprocess
import shutil
from datetime import datetime
from pathlib import Path
import openpyxl

# ──────────────────────────────────────────
# 設定
# ──────────────────────────────────────────
ROOT       = Path(__file__).parent.parent
HTML_SRC   = ROOT / "経営ダッシュボード_糸貫グループ (1) (1) (4).html"
HTML_DEST  = ROOT / "itanuki_dashboard.html"   # URLに使うクリーンな名前
EXCEL_PATH = ROOT / "09_system" / "月次データ入力テンプレート.xlsx"
BACKUP_DIR = ROOT / "09_system" / "dashboard_backups"

# GitHub Pages URL（ユーザー名・リポジトリ名から自動生成）
GITHUB_PAGES_URL = "https://saepin0519.github.io/ai-agent/itanuki_dashboard.html"

DIVISIONS_METRICS = {
    "糸貫工場": ["売上","原価","粗利","営業利益",
                 "基本給","時間外","雑給","賞与引当","法定福利","退職金",
                 "消耗品","修繕費","地代家賃","減価償却","電力","燃料"],
    "加工":     ["売上","原価","粗利","営業利益",
                 "基本給","時間外","雑給","賞与引当","法定福利","退職金",
                 "消耗品","減価償却","電力","燃料"],
    "糸貫検査": ["売上","原価","粗利","営業利益",
                 "基本給","時間外","雑給","賞与引当","法定福利","退職金",
                 "消耗品","減価償却","電力","燃料"],
    "本社検査": ["売上","原価","粗利","営業利益",
                 "基本給","時間外","雑給","賞与引当","法定福利","退職金",
                 "消耗品","減価償却","電力","燃料"],
    "物流検査": ["売上","原価","粗利","営業利益",
                 "基本給","時間外","雑給","賞与引当","法定福利","退職金",
                 "消耗品","減価償却","電力","燃料"],
}


# ──────────────────────────────────────────
# Excel 読み込み
# ──────────────────────────────────────────
def read_excel() -> dict:
    """
    Excelから月ラベルと各部門データを読み込む。
    戻り値: {
        'new_m':       'R8.2月',
        'new_ms':      '2月',
        'latest_label':'R8年2月',
        'data': { '糸貫工場': { '売上': 195000000, ... }, ... }
    }
    """
    if not EXCEL_PATH.exists():
        print(f"エラー: Excelファイルが見つかりません")
        print(f"  {EXCEL_PATH}")
        print("\n先に以下を実行してテンプレートを作成してください:")
        print("  python 09_system/create_excel_template.py")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["データ入力"]

    # 月ラベル (B3, B4, B5)
    new_m        = str(ws["B3"].value or "").strip()
    new_ms       = str(ws["B4"].value or "").strip()
    latest_label = str(ws["B5"].value or "").strip()

    if not new_m or not new_ms:
        print("エラー: Excelの「新しい月」ラベルが入力されていません。")
        print("  B3（長ラベル）と B4（短ラベル）を入力してください。")
        sys.exit(1)

    if not latest_label:
        latest_label = new_m

    # データ読み込み
    data = {div: {} for div in DIVISIONS_METRICS}
    current_div = None

    for row in ws.iter_rows(min_row=8, values_only=True):
        a, b, c = (row[0] or ""), (row[1] or ""), (row[2] or 0)

        # 部門ヘッダー行（例: "▼ 糸貫工場"）
        a_str = str(a).replace("▼", "").strip()
        if a_str in DIVISIONS_METRICS:
            current_div = a_str
            continue

        # データ行
        if current_div and str(b).strip():
            metric = str(b).strip()
            try:
                amount = int(float(str(c).replace(",", "") or "0"))
            except (ValueError, TypeError):
                amount = 0
            data[current_div][metric] = amount

    return {
        "new_m": new_m,
        "new_ms": new_ms,
        "latest_label": latest_label,
        "data": data,
    }


# ──────────────────────────────────────────
# HTML 更新ロジック
# ──────────────────────────────────────────
def parse_month_arrays(html: str):
    m_match  = re.search(r"const M\s*=\s*\[([^\]]+)\]", html)
    ms_match = re.search(r"const Ms\s*=\s*\[([^\]]+)\]", html)
    if not m_match or not ms_match:
        print("エラー: HTMLの月配列が見つかりません")
        sys.exit(1)
    parse = lambda raw: [s.strip().strip("'\"") for s in raw.split(",")]
    return parse(m_match.group(1)), parse(ms_match.group(1))


def parse_metric_array(html: str, div: str, metric: str) -> list[int]:
    pattern = rf"{re.escape(div)}:\s*\{{(.*?)\}},"
    div_m = re.search(pattern, html, re.DOTALL)
    if not div_m:
        return []
    mp = re.search(rf"{re.escape(metric)}:\s*\[([^\]]+)\]", div_m.group(1))
    if not mp:
        return []
    return [int(v.strip()) for v in mp.group(1).split(",")]


def update_month_arrays(html, old_m, old_ms, new_m, new_ms):
    updated_m  = old_m[1:]  + [new_m]
    updated_ms = old_ms[1:] + [new_ms]
    html = re.sub(r"const M\s*=\s*\[[^\]]+\]",
                  f"const M = [{', '.join(repr(x) for x in updated_m)}]", html)
    html = re.sub(r"const Ms\s*=\s*\[[^\]]+\]",
                  f"const Ms = [{', '.join(repr(x) for x in updated_ms)}]", html)
    return html, updated_m


def update_metric_in_html(html, div, metric, old_vals, new_val):
    updated = old_vals[1:] + [new_val]
    vals_str = ", ".join(f"{v:>12}" for v in updated)
    div_pattern = rf"({re.escape(div)}:\s*\{{)(.*?)(\}},)"
    def replacer(m):
        block = re.sub(
            rf"({re.escape(metric)}:\s*\[)[^\]]+(\])",
            rf"\g<1>{vals_str}\2",
            m.group(2)
        )
        return m.group(1) + block + m.group(3)
    return re.sub(div_pattern, replacer, html, flags=re.DOTALL)


def update_header(html, m_list, latest_label):
    html = re.sub(
        r"糸貫工場 / 加工 / 糸貫検査 / 本社検査 / 物流検査｜[^<]+",
        f"糸貫工場 / 加工 / 糸貫検査 / 本社検査 / 物流検査｜{m_list[0]}〜{m_list[-1]}",
        html
    )
    html = re.sub(
        r"<strong>最新月：[^<]+</strong>",
        f"<strong>最新月：{latest_label}</strong>",
        html
    )
    return html


def backup(html_content: str):
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = BACKUP_DIR / f"backup_{ts}.html"
    path.write_text(html_content, encoding="utf-8")
    print(f"  バックアップ: dashboard_backups/backup_{ts}.html")


# ──────────────────────────────────────────
# Git Push
# ──────────────────────────────────────────
def git_push(month_label: str):
    """HTMLをコミットしてGitHubにPush"""
    try:
        subprocess.run(["git", "-C", str(ROOT), "add",
                        "itanuki_dashboard.html"], check=True)
        subprocess.run(["git", "-C", str(ROOT), "commit", "-m",
                        f"ダッシュボード更新: {month_label}"], check=True)
        subprocess.run(["git", "-C", str(ROOT), "push"], check=True)
        return True
    except subprocess.CalledProcessError as e:
        print(f"\n⚠️  Git Push に失敗しました: {e}")
        print("   手動で push してください:")
        print("   git add itanuki_dashboard.html && git commit -m '更新' && git push")
        return False


# ──────────────────────────────────────────
# メイン処理
# ──────────────────────────────────────────
def main():
    print("=" * 60)
    print("  糸貫グループ 月次ダッシュボード更新")
    print("=" * 60)

    # 1. Excelから読み込み
    print("\n📊 Excelファイルを読み込んでいます...")
    excel_data = read_excel()
    new_m        = excel_data["new_m"]
    new_ms       = excel_data["new_ms"]
    latest_label = excel_data["latest_label"]
    div_data     = excel_data["data"]

    print(f"  新しい月: {new_m}（表示: {latest_label}）")
    for div in DIVISIONS_METRICS:
        total_sales = div_data[div].get("売上", 0)
        print(f"  {div}: 売上 {total_sales:,}円")

    print()
    confirm = input("この内容で更新しますか？ [y/N]: ").strip().lower()
    if confirm != "y":
        print("キャンセルしました。")
        sys.exit(0)

    # 2. HTMLを読み込み・バックアップ
    print("\n💾 バックアップを作成中...")
    html = HTML_DEST.read_text(encoding="utf-8") if HTML_DEST.exists() \
           else HTML_SRC.read_text(encoding="utf-8")
    backup(html)

    # 3. 月配列を更新
    old_m, old_ms = parse_month_arrays(html)
    print(f"\n📅 期間更新: {old_m[0]} 〜 {old_m[-1]}  →  {old_m[1]} 〜 {new_m}")
    html, new_m_list = update_month_arrays(html, old_m, old_ms, new_m, new_ms)

    # 4. 各部門データを更新
    print("📝 データを書き込み中...")
    for div, metrics in DIVISIONS_METRICS.items():
        for metric in metrics:
            old_vals = parse_metric_array(html, div, metric)
            if not old_vals:
                continue
            new_val = div_data[div].get(metric, 0)
            html = update_metric_in_html(html, div, metric, old_vals, new_val)

    # 5. ヘッダー更新
    html = update_header(html, new_m_list, latest_label)

    # 6. 保存
    HTML_DEST.write_text(html, encoding="utf-8")
    print(f"✅ HTML 保存完了: {HTML_DEST.name}")

    # 7. GitHub Push
    print("\n🚀 GitHub へ Push 中...")
    success = git_push(new_m)

    # 8. URL表示
    print("\n" + "=" * 60)
    if success:
        print("  ✅ 更新・公開が完了しました！")
        print()
        print("  📎 上司への共有URL:")
        print(f"  {GITHUB_PAGES_URL}")
        print()
        print("  ※ GitHub Pages が有効な場合、数分で反映されます")
    else:
        print("  HTML は更新済みです。手動で push してください。")
    print("=" * 60)


if __name__ == "__main__":
    main()
