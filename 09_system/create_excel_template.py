"""
糸貫グループ 月次データ入力 Excelテンプレート生成スクリプト
==============================================================
最初に1回だけ実行してください。Excelテンプレートを作成します。

使い方:
    pip install openpyxl
    python 09_system/create_excel_template.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "09_system" / "月次データ入力テンプレート.xlsx"

# ──────────────────────────────────────────
# データ定義
# ──────────────────────────────────────────
DIVISIONS_METRICS = [
    ("糸貫工場", [
        "売上", "原価", "粗利", "営業利益",
        "基本給", "時間外", "雑給", "賞与引当", "法定福利", "退職金",
        "消耗品", "修繕費", "地代家賃", "減価償却", "電力", "燃料",
    ]),
    ("加工", [
        "売上", "原価", "粗利", "営業利益",
        "基本給", "時間外", "雑給", "賞与引当", "法定福利", "退職金",
        "消耗品", "減価償却", "電力", "燃料",
    ]),
    ("糸貫検査", [
        "売上", "原価", "粗利", "営業利益",
        "基本給", "時間外", "雑給", "賞与引当", "法定福利", "退職金",
        "消耗品", "減価償却", "電力", "燃料",
    ]),
    ("本社検査", [
        "売上", "原価", "粗利", "営業利益",
        "基本給", "時間外", "雑給", "賞与引当", "法定福利", "退職金",
        "消耗品", "減価償却", "電力", "燃料",
    ]),
    ("物流検査", [
        "売上", "原価", "粗利", "営業利益",
        "基本給", "時間外", "雑給", "賞与引当", "法定福利", "退職金",
        "消耗品", "減価償却", "電力", "燃料",
    ]),
]

# ──────────────────────────────────────────
# スタイル定義
# ──────────────────────────────────────────
NAVY   = "17305A"
BLUE   = "2E75B6"
LBLUE  = "BDD7EE"
LGRAY  = "F2F2F2"
GREEN  = "375623"
LGREEN = "E2EFDA"
WHITE  = "FFFFFF"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(thin=True):
    s = Side(style="thin" if thin else "medium", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, color=WHITE, bg=None,
             align="left", number_format=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="游ゴシック", size=10, bold=bold,
                     color="000000" if bg is None else color)
    if bg:
        cell.fill = make_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=False)
    if border:
        cell.border = make_border()
    if number_format:
        cell.number_format = number_format
    return cell


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "データ入力"

    # ── 列幅 ──
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40

    # ── タイトル ──
    ws.row_dimensions[1].height = 30
    title_cell = ws.cell(row=1, column=1,
                         value="糸貫グループ  月次損益データ入力シート")
    title_cell.font = Font(name="游ゴシック", size=14, bold=True, color=WHITE)
    title_cell.fill = make_fill(NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")

    # ── 月ラベル入力エリア ──
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22

    label_data = [
        (3, "新しい月（長）", "例: R8.2月",  "月ラベル（グラフの凡例・ヘッダーに使用）"),
        (4, "新しい月（短）", "例: 2月",     "グラフの横軸ラベル"),
        (5, "ヘッダー表示",   "例: R8年2月",  "ダッシュボード右上の「最新月」表示"),
    ]
    for row, label, example, note in label_data:
        set_cell(ws, row, 1, label, bold=True, bg=BLUE, color=WHITE, align="center")
        set_cell(ws, row, 2, "",    bg=LGREEN)   # 入力セル
        set_cell(ws, row, 3, example, bg=LGRAY, align="center")
        c = ws.cell(row=row, column=4, value=note)
        c.font = Font(name="游ゴシック", size=9, color="808080", italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ── データ表ヘッダー ──
    ws.row_dimensions[7].height = 22
    headers = [("A7", "部門"), ("B7", "科目"), ("C7", "金額（円）"), ("D7", "備考")]
    for addr, val in headers:
        r, c_str = int(addr[1:]), addr[0]
        c = ord(c_str) - ord("A") + 1
        set_cell(ws, r, c, val, bold=True, bg=NAVY, color=WHITE, align="center")

    # ── データ行 ──
    row = 8
    div_colors = {
        "糸貫工場": (BLUE,   LBLUE),
        "加工":     ("4472C4", "DCE6F1"),
        "糸貫検査": ("7030A0", "E6D9F0"),
        "本社検査": ("C55A11", "FCE4D6"),
        "物流検査": ("375623", LGREEN),
    }

    for div, metrics in DIVISIONS_METRICS:
        hdr_bg, row_bg = div_colors.get(div, (BLUE, LBLUE))

        # 部門ヘッダー行
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=1, value=f"▼ {div}")
        c.font = Font(name="游ゴシック", size=10, bold=True, color=WHITE)
        c.fill = make_fill(hdr_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = make_border()
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        for metric in metrics:
            ws.row_dimensions[row].height = 19
            is_profit = metric in ("粗利", "営業利益")
            bg = LGRAY if is_profit else row_bg

            set_cell(ws, row, 1, div,    bg=LGRAY, align="center")
            set_cell(ws, row, 2, metric, bg=bg, bold=is_profit)
            # 金額入力セル（黄色背景）
            input_cell = ws.cell(row=row, column=3, value=0)
            input_cell.font = Font(name="游ゴシック", size=10,
                                   color="C00000" if is_profit else "000000")
            input_cell.fill = make_fill("FFFF99")
            input_cell.alignment = Alignment(horizontal="right", vertical="center")
            input_cell.number_format = '#,##0'
            input_cell.border = make_border()
            # 備考列（空白）
            set_cell(ws, row, 4, "", bg=WHITE)
            row += 1

        row += 1  # 部門間の空行

    # ── 使い方説明シート ──
    ws2 = wb.create_sheet("使い方")
    guide = [
        ("A1", "【使い方ガイド】", True, NAVY, WHITE),
        ("A3", "① 「データ入力」シートを開く", True, None, "000000"),
        ("A4", "② 上部の「新しい月」ラベルを3行とも入力する", False, None, "000000"),
        ("A5", "③ 各部門の黄色いセル（金額欄）に数値を入力する（単位：円）", False, None, "000000"),
        ("A6", "④ Excelファイルを上書き保存する", False, None, "000000"),
        ("A7", "⑤ コマンドプロンプトで以下を実行：", False, None, "000000"),
        ("A8", "   python 09_system/dashboard_monthly_update.py", False, LGRAY, "000000"),
        ("A10", "⑥ 自動でHTMLが更新され、GitHubにアップロードされます", False, None, "000000"),
        ("A11", "⑦ 発行されたURLを上司に共有してください", False, None, "000000"),
        ("A13", "【注意事項】", True, None, "C00000"),
        ("A14", "・粗利・営業利益は自動計算が理想ですが、手入力でも構いません", False, None, "000000"),
        ("A15", "・0円の項目は0のまま（空欄にしない）", False, None, "000000"),
        ("A16", "・前月データのバックアップは自動作成されます", False, None, "000000"),
    ]
    for addr, text, bold, bg, fg in guide:
        c = ws2[addr]
        c.value = text
        c.font = Font(name="游ゴシック", size=10, bold=bold, color=fg)
        if bg:
            c.fill = make_fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.column_dimensions["A"].width = 70

    wb.save(OUTPUT_PATH)
    print(f"✅ テンプレートを作成しました:")
    print(f"   {OUTPUT_PATH}")


if __name__ == "__main__":
    create_template()
